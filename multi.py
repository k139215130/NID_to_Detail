# -*- coding: utf-8 -*-
import getpass
import requests
from bs4 import BeautifulSoup
import re
import os
from openpyxl import load_workbook, Workbook

# 讀檔後放入 nidList
nidList = []
wb = load_workbook('example.xlsx', read_only=True)
sheet = wb.active
for row in sheet.rows:
	for cell in row:
		nidList.append(cell.value)

print('讀檔完畢!')

# 獲取登入網頁
s = requests.Session()
try:
	login = s.get("http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
except:
	print('未使用校內網路')
	os._exit(0)

# 登入
username = input('請輸入學號:')
password = getpass.getpass('請輸入密碼:')

loginData = {
	'asn_code': 'A66',  # 社團代號
	'auserid': username,  # 帳號
	'apwd': password  # 密碼
}

login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)

bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面

if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
	print('密碼輸入錯誤!')
	os._exit(0)
elif re.search('對不起!!您無權進入系統!', bs_loginPage):
	print('帳號密碼輸入錯誤!')
	os._exit(0)

print('登入成功!')

wb = Workbook(write_only=True)
ws = wb.create_sheet('詳細資料')

print('開始查詢!')

for nid in nidList:
	searchData = {
		'stuid': nid, # 學號
		'idButton': '送出'
	}

	searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)

	bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面

	if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
		d = {'學號': nid, '系級': '查無資料', '姓名': '查無資料','性別': '查無資料', '出生年月日': '查無資料'}
	else:
		result = bs_searchPage.select("table.tableStyle td.tableContentLeft")
		newResult = []

		for i in result:
			newResult.append(re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))

		d = {'學號': newResult[0], '系級': newResult[1], '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}

	ws.append([d['學號'], d['系級'], d['姓名'], d['性別'], d['出生年月日']])

print('查詢完畢!!!')
print('開始寫檔!!!')

wb.save('output.xlsx') 

print('寫檔完畢!!!')