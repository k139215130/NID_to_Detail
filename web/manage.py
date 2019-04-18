#!/usr/bin/env python3
from flask import Flask, render_template, flash, send_file
from flask_bootstrap import Bootstrap
from flask_sqlalchemy import SQLAlchemy

#Form
from forms import SingleForm, MultiForm

# get nid
import requests
from bs4 import BeautifulSoup
import re

# excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from tempfile import NamedTemporaryFile
from io import BytesIO
import os

# chart
from pyecharts import Line, Bar, Pie

app = Flask(__name__)
app.config["SECRET_KEY"] = "hard to guess string"
app.config['UPLOAD_FOLDER'] = "hard to guess string"
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///test.db'
bootstrap = Bootstrap(app)
db = SQLAlchemy(app)

"""Models"""
class Activity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)

    #全部社課
    @classmethod
    def get_all_activity_list(cls):
        return cls.query.all()

class Member(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    nid = db.Column(db.String(10), nullable=False)
    department = db.Column(db.String(10), nullable=False)
    # department_level = db.Column(db.String(10), nullable=False)
    # department = db.Column(db.String(30), nullable=False)
    # level = db.Column(db.String(10), nullable=False)
    # college = db.Column(db.String(10), nullable=False)
    sex = db.Column(db.String(1), nullable=False)
    birthday = db.Column(db.String(10), nullable=False)
    activity_id = db.Column(db.Integer, db.ForeignKey('activity.id'),nullable=False)
    activity = db.relationship('Activity',backref=db.backref('members', lazy=True))

    # def __init__(self, **kwargs):
    #     super(Member, self).__init__(**kwargs)
    #     if re.search(r'一', i.department_level):
    #         self.level = '大一'
    #     elif re.search(r'二', i.department):
    #         self.level = '大二'
    #     elif re.search(r'三', i.department):
    #         self.level = '大三'
    #     elif re.search(r'四', i.department):
    #         self.level = '大四'

    #單一社課人數總和
    @classmethod
    def get_activity_member_count(cls, id):
        return cls.query.filter_by(activity_id=id).count()
    #單一社課全部資料
    @classmethod
    def get_activity_member(cls, id):
        return cls.query.filter_by(activity_id=id).all()
    #單一社課男女人數
    @classmethod
    def get_sex_number(cls, id):
        boy = cls.query.filter_by(activity_id=id, sex='男').count()
        girl = cls.query.filter_by(activity_id=id, sex='女').count()
        return [boy,girl]
    #單一社課年級人數
    @classmethod
    def get_level_number(cls, id):
        result = [0,0,0,0]
        for i in cls.query.filter_by(activity_id=id).all():
            if re.search(r'一', i.department):
                result[0] = result[0] + 1
            elif re.search(r'二', i.department):
                result[1] = result[1] + 1
            elif re.search(r'三', i.department):
                result[2] = result[2] + 1
            elif re.search(r'四', i.department):
                result[3] = result[3] + 1
        return result
    #單一社課各系人數
    @classmethod
    def get_department_number(cls, id):
        result = {}
        for i in cls.query.filter_by(activity_id=id).all():
            s = re.findall(r'[^一,二,三,四]*', i.department)[0]
            print(s)
            try:
                result[s] = result[s]+1
            except KeyError:
                result[s] = 1
        return result

@app.route("/")
def index():
    return render_template('index.html')

@app.route("/single", methods=['GET','POST'])
def single():
    form = SingleForm()
    if form.validate_on_submit():
        s = requests.Session()
        # 檢查網路狀態
        try:
          login = s.get("http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            flash('未使用校內網路', 'error')
        loginData = {
          'asn_code': 'A66',  # 社團代號
          'auserid': form.username.data,  # 帳號
          'apwd': form.password.data # 密碼
        }
        login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面
        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            flash('密碼輸入錯誤!', 'error')
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            flash('帳號密碼輸入錯誤!', 'error')
        searchData = {
          'stuid': form.nid.data,  # 學號
          'idButton': '送出'
        }
        searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
        bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面
        if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
            result = {'學號': form.nid.data, '系級': '查無資料', '姓名': '查無資料', '性別': '查無資料', '出生年月日': '查無資料'}
        else:
            result = bs_searchPage.select("table.tableStyle td.tableContentLeft")
            newResult = []
            for i in result:
                newResult.append(re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
            result = {'學號': newResult[0], '系級': newResult[1], '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}
        return render_template('single.html', form=form, result=result)
    return render_template('single.html', form=form)

@app.route("/chart", methods=['GET', 'POST'])
def chart():
    all_course_people = [] #全部社課人數
    all_course_detail = [] #全部社課詳細資料
    all_course_name = [] #全部社課名稱
    sex_number_charts = [] #全部社課男女人數圖表
    level_number_charts = [] #全部社課年級人數圖表
    department_number_charts = [] #全部社課各系人數

    for i in Activity.get_all_activity_list():
        all_course_name.append(i.name)
        all_course_people.append(Member.get_activity_member_count(id=i.id))
        all_course_detail.append(Member.get_activity_member(id=i.id))

        #單一社課男女人數(圖)
        pie = Pie(i.name, "男女人數")
        pie.add("人數", ['男生', '女生'], Member.get_sex_number(id=i.id))
        sex_number_charts.append(pie.render_embed())
        #單一社課年級人數(圖)
        bar = Bar(i.name, "年級人數")
        bar.add("人數", ['大一', '大二', '大三', '大四'], Member.get_level_number(id=i.id))
        level_number_charts.append(bar.render_embed())
        #單一社課各系人數
        result = Member.get_department_number(id=i.id)
        department_list = []
        people_list = []
        for d in result:
            department_list.append(d)
            people_list.append(result[d])
        bar = Bar(i.name, "各系人數")
        bar.add("人數", department_list, people_list)
        department_number_charts.append(bar.render_embed())
        del result, department_list, people_list
        #單一社課各學院人數
    
    #全部社課總人數+名稱(直線圖)
    line = Line("黑客社", "社課總人數")
    line.add("人數", all_course_name, all_course_people)
    all_course_charts = line.render_embed()

    #bar.add("人數", ["B", "C", "D", "E", "F", "G"], [5, 20, 36, 10, 75, 90])
    return render_template('chart.html', all_course_charts=all_course_charts, all_course_name=all_course_name, all_course_detail=all_course_detail, sex_number_charts=sex_number_charts, level_number_charts=level_number_charts, department_number_charts=department_number_charts)

@app.route("/multi", methods=['GET','POST'])
def multi():
    form = MultiForm()
    if form.validate_on_submit():
        if form.store.data:
            act = Activity(name=form.name.data)
            db.session.add(act)
        f = form.file.data
        f.save(f.filename)
        # 讀檔後放入 nidList
        nidList = []
        wb = load_workbook(f.filename, read_only=True)
        sheet = wb.active
        for row in sheet.rows:
            for cell in row:
                nidList.append(cell.value)
        # 獲取登入網頁
        s = requests.Session()
        try:
            login = s.get("http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            flash('未使用校內網路', 'error')
        # 登入
        loginData = {
            'asn_code': 'A66',  # 社團代號
            'auserid': form.username.data,  # 帳號
            'apwd': form.password.data  # 密碼
        }
        login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面
        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            flash('密碼輸入錯誤!', 'error')
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            flash('帳號密碼輸入錯誤!', 'error')
        wb = Workbook()
        ws = wb.create_sheet('詳細資料', 0)
        for nid in nidList:
            searchData = {
                'stuid': nid, # 學號
                'idButton': '送出'
            }
            searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
            bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面
            if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
                d = {'學號': nid, '系級': '查無資料', '姓名': '查無資料','性別': '查無資料', '出生年月日': '查無資料'}
            else:
                result = bs_searchPage.select("table.tableStyle td.tableContentLeft")
                newResult = []
                for i in result:
                    newResult.append(re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
                d = {'學號': newResult[0], '系級': newResult[1], '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}
            if form.store.data:
                member = Member(name=d['姓名'], nid=d['學號'], department_level=d['系級'], sex=d['性別'], birthday=d['出生年月日'], activity=act)
                db.session.add(member)
            ws.append([d['學號'], d['系級'], d['姓名'], d['性別'], d['出生年月日']])
        db.session.commit()
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            return send_file(BytesIO(stream), attachment_filename='output.xlsx', as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template('multi.html', form=form)

@app.route("/")
def show():
    return render_template('show.html')

if __name__ == '__main__':
    app.run(debug=True)
