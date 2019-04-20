from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from .models import Activity, Member

#Get NID
import requests
from bs4 import BeautifulSoup
import re

#Read Excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
import os
from openpyxl.writer.excel import save_virtual_workbook

def index(request):
    return render(request, 'index.html', {})


def single(request):
    result = {}
    if request.method == "POST":
        s = requests.Session()
        # 檢查網路狀態
        try:
            login = s.get(
                "http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            messages.add_message(request, messages.ERROR, '未使用校內網路!', extra_tags='all')
            return render(request, 'single.html', {})
        loginData = {
            'asn_code': 'A66',  # 社團代號
            'auserid': request.POST.get('username'),  # 帳號
            'apwd': request.POST.get('password')  # 密碼
        }
        login = s.post(
            'http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(
            login.text, "html.parser").get_text(strip=True)  # 分析登入頁面
        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'single.html', {})
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'single.html', {})
        searchData = {
            'stuid': request.POST.get('nid'),  # 學號
            'idButton': '送出'
        }
        searchPage = s.post(
            'http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
        bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面
        if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
            result = {'學號': request.POST.get('nid'), '系級': '查無資料',
                      '姓名': '查無資料', '性別': '查無資料', '出生年月日': '查無資料'}
        else:
            result = bs_searchPage.select(
                "table.tableStyle td.tableContentLeft")
            newResult = []
            for i in result:
                newResult.append(
                    re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
            result = {'學號': newResult[0], '系級': newResult[1],
                      '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}
        return render(request, 'single.html', {'result':result})
    return render(request, 'single.html', {})


def multi(request):
    if request.method == "POST":
        print(request.POST.get('check')) #None on
        # 讀檔後放入 nidList
        nidList = []
        wb = load_workbook(filename=request.FILES['file'], read_only=True)
        sheet = wb.active
        for row in sheet.rows:
            for cell in row:
                nidList.append(cell.value)
        # 獲取登入網頁
        s = requests.Session()
        try:
            login = s.get("http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            messages.add_message(request, messages.ERROR, '未使用校內網路!', extra_tags='all')
            return render(request, 'multi.html', {})
        # 登入
        loginData = {
            'asn_code': 'A66',  # 社團代號
            'auserid': request.POST.get('username'),  # 帳號
            'apwd': request.POST.get('password')  # 密碼
        }
        login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面
        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'multi.html', {})
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'multi.html', {})
        wb = Workbook()
        ws = wb.create_sheet('詳細資料', 0)
        for nid in nidList:
            searchData = {
                'stuid': nid, # 學號
                'idButton': '送出'
            }
            searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
            bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面
            if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
                d = {'學號': nid, '系級': '查無資料', '姓名': '查無資料','性別': '查無資料', '出生年月日': '查無資料'}
            else:
                result = bs_searchPage.select("table.tableStyle td.tableContentLeft")
                newResult = []
                for i in result:
                    newResult.append(re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
                d = {'學號': newResult[0], '系級': newResult[1], '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}
            # if form.store.data:
            #     member = Member(name=d['姓名'], nid=d['學號'], department=d['系級'], sex=d['性別'], birthday=d['出生年月日'], activity=act)
            #     db.session.add(member)
            ws.append([d['學號'], d['系級'], d['姓名'], d['性別'], d['出生年月日']])
        # db.session.commit()
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=Output.xlsx'
        return response
    return render(request, 'multi.html', {})


def chart(request):
    return render(request, 'chart.html', {})


def edit(request):
    return render(request, 'edit.html', {})