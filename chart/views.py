from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

#Get NID
import requests
from bs4 import BeautifulSoup
import re

#Read Excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
import os
from openpyxl.writer.excel import save_virtual_workbook

#Chart
from pyecharts import Line, Bar, Pie

# forms
from .forms import SingleForm, ActivityTagForm

# model
from .models import ActivityTag, Activity, Member

def single(request):
    form = SingleForm()
    result = {}
    if request.method == "POST":
        form = SingleForm(request.POST)
        s = requests.Session()
        # 檢查網路狀態
        try:
            login = s.get(
                "http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            messages.add_message(request, messages.ERROR, '未使用校內網路!')
            form = SingleForm()
            result = {}
            return render(request, 'single.html', {'form':form, 'result':result})
        
        loginData = {
            'asn_code': 'A66',  # 社團代號
            'auserid': request.POST.get('username'),  # 帳號
            'apwd': request.POST.get('password')  # 密碼
        }
        login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面

        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!')
            form = SingleForm()
            result = {}
            return render(request, 'single.html', {'form':form, 'result':result})
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!')
            form = SingleForm()
            result = {}
            return render(request, 'single.html', {'form':form, 'result':result})
        
        searchData = {
            'stuid': request.POST.get('nid'),  # 學號
            'idButton': '送出'
        }
        searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
        bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面

        if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
            result = {'學號': request.POST.get('nid'), '系級': '查無資料',
                      '姓名': '查無資料', '性別': '查無資料', '出生年月日': '查無資料'}
        else:
            result = bs_searchPage.select("table.tableStyle td.tableContentLeft")
            newResult = []
            for i in result:
                newResult.append(
                    re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
            result = {'學號': newResult[0], '系級': newResult[1],
                      '姓名': newResult[2], '性別': newResult[3], '出生年月日': newResult[4]}
        form = SingleForm()
    return render(request, 'single.html', {'form':form, 'result':result})


@login_required
@csrf_exempt
def multi(request):
    tag = ActivityTag.objects.all()
    addform = ActivityTagForm()
    
    # Add
    if request.method == "POST" and 'create' in request.POST:
        addform = ActivityTagForm(request.POST)
        if addform.is_valid():
            addform.save(commit=True)
            addform = ActivityTagForm()
            tag = ActivityTag.objects.all()
        return render(request, 'multi.html', {'tag':tag, 'addform':addform})
        
    # Mutil
    if request.is_ajax():
        # 讀檔後放入 nidList
        nidList = []
        wb = load_workbook(filename=request.FILES['file'], read_only=True)
        sheet = wb.active
        for row in sheet.rows:
            for cell in row:
                nidList.append(cell.value)
        # 獲取登入網頁
        s = requests.Session()
        try:
            login = s.get("http://infocenter.fcu.edu.tw/assoc/assoc_login.jsp", timeout=3)
        except:
            messages.add_message(request, messages.ERROR, '未使用校內網路!', extra_tags='all')
            return render(request, 'multi.html', {})
        # 登入
        loginData = {
            'asn_code': 'A66',  # 社團代號
            'auserid': request.POST.get('username'),  # 帳號
            'apwd': request.POST.get('password')  # 密碼
        }
        login = s.post('http://infocenter.fcu.edu.tw/assoc/authenticate.jsp', data=loginData)
        bs_loginPage = BeautifulSoup(login.text, "html.parser").get_text(strip=True)  # 分析登入頁面
        if re.search('對不起!!您的帳號/密碼有誤!!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'multi.html', {})
        elif re.search('對不起!!您無權進入系統!', bs_loginPage):
            messages.add_message(request, messages.ERROR, '帳號密碼輸入錯誤!', extra_tags='user')
            return render(request, 'multi.html', {})
        result = []
        for nid in nidList:
            searchData = {
                'stuid': nid, # 學號
                'idButton': '送出'
            }
            searchPage = s.post('http://infocenter.fcu.edu.tw/assoc/assoc30.jsp', data=searchData)
            bs_searchPage = BeautifulSoup(searchPage.text, "html.parser")  # 分析查詢頁面
            if re.search('抱歉, 資料不存在!', bs_searchPage.get_text(strip=True)):
                d = {'學號': nid, '系級': '查無資料', '姓名': '查無資料','性別': '查無資料', '出生年月日': '查無資料'}
                result.append(d)
            else:
                page = bs_searchPage.select("table.tableStyle td.tableContentLeft")
                tmpResult = []
                for i in page:
                    tmpResult.append(re.search(r'>(.*)<', re.sub(r'\s', '', str(i))).group(1))
                d = {'學號': tmpResult[0], '系級': tmpResult[1], '姓名': tmpResult[2], '性別': tmpResult[3], '出生年月日': tmpResult[4]}
                result.append(d)
            
            # 產生 Excel
            wb = Workbook()
            ws = wb.create_sheet('詳細資料', 0)
            for i in result:
                ws.append(i)

            # 儲存置資料庫
            if request.POST.get('check') == 'on':
                tagTmp = ActivityTag.objects.get(name=request.POST.get('tag'))
                ActivityTmp = tagTmp.tags.create(name=request.POST.get('activityname'))
                for i in result:
                    ActivityTmp.activitys.create(name=i[0], nid=i[1], department=i[2], sex=i[3], birthday=i[4])
           
        return HttpResponse(save_virtual_workbook(wb))
    return render(request, 'multi.html', {'tag':tag, 'addform':addform})


def chart(request):
    all_course_people = [] #全部社課人數
    all_course_detail = [] #全部社課詳細資料
    all_course_name = [] #全部社課名稱
    sex_number_charts = [] #全部社課男女人數圖表
    level_number_charts = [] #全部社課年級人數圖表
    department_number_charts = [] #全部社課各系人數

    for i in Activity.get_all_activity_list():
        all_course_name.append(i.name)
        all_course_people.append(Member.get_activity_member_count(act=i))
        all_course_detail.append(Member.get_activity_member(act=i))
        #單一社課男女人數(圖)
        pie = Pie(i.name, "男女人數")
        pie.add("人數", ['男生', '女生'], Member.get_sex_number(act=i))
        sex_number_charts.append(pie.render_embed())
        #單一社課年級人數(圖)
        bar = Bar(i.name, "年級人數")
        bar.add("人數", ['大一', '大二', '大三', '大四'], Member.get_level_number(act=i))
        level_number_charts.append(bar.render_embed())
        #單一社課各系人數
        result = Member.get_department_number(act=i)
        department_list = []
        people_list = []
        for d in result:
            department_list.append(d)
            people_list.append(result[d])
        bar = Bar(i.name, "各系人數")
        bar.add("人數", department_list, people_list)
        department_number_charts.append(bar.render_embed())
        del result, department_list, people_list
        #單一社課各學院人數
        """未完成"""

    #全部社課總人數+名稱(直線圖)
    line = Line("黑客社", "社課總人數")
    line.add("人數", all_course_name, all_course_people)
    all_course_charts = line.render_embed()
    return render(request, 'chart.html', { 'all_course_charts':all_course_charts, 'all_course_name':all_course_name, 'all_course_detail':all_course_detail, 'sex_number_charts':sex_number_charts, 'level_number_charts':level_number_charts, 'department_number_charts':department_number_charts})


@login_required
def edit(request):
    return render(request, 'edit.html', {})